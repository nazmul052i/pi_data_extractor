import csv
import pandas as pd


class DataExporter:
    """Handles CSV, TSV, TXT (with instrument tag replacement), IQ, and XLSX export formats"""
    
    def __init__(self, dataframe, descriptions=None, units=None, timezone="Local", 
                 instrument_mapping=None):
        self.dataframe = dataframe
        self.descriptions = descriptions or {}
        self.units = units or {}
        self.timezone = timezone
        self.instrument_mapping = instrument_mapping or {}  # Maps original tags to instrument paths
    
    def get_clean_dataframe(self):
        """Get dataframe without status columns"""
        clean_df = self.dataframe[[col for col in self.dataframe.columns 
                                if col != "Status"]].copy()
        
        # Remove timezone from timestamp for compatibility
        if "Timestamp" in clean_df.columns and pd.api.types.is_datetime64_any_dtype(clean_df["Timestamp"]):
            clean_df["Timestamp"] = pd.to_datetime(clean_df["Timestamp"]).dt.tz_localize(None)
        
        return clean_df
    
    def parse_instrument_tag_from_opc_path(self, opc_path):
        """
        Parse instrument tag from OPC path for .txt export
        Examples:
        - 'E20FC0023/PID1/PV.CV' → 'E20FC0023.PV'
        - 'UNIT1/TANK101/LEVEL.CV' → 'TANK101.LEVEL'
        - 'FIC201A/OUT.CV' → 'FIC201A.OUT'
        """
        if not opc_path or not opc_path.strip():
            return ''
        
        import re
        
        # Clean the path
        cleaned_path = opc_path.strip()
        
        # Strategy 1: Handle common OPC patterns with / separator
        if '/' in cleaned_path:
            parts = cleaned_path.split('/')
            
            # Pattern: INSTRUMENT/MODULE/SIGNAL.CV → INSTRUMENT.SIGNAL
            if len(parts) >= 2:
                instrument_part = parts[0]  # E20FC0023
                signal_part = parts[-1]     # PV.CV
                
                # Extract signal name from signal_part (remove .CV suffix)
                if '.' in signal_part:
                    signal_name = signal_part.split('.')[0]  # PV
                else:
                    signal_name = signal_part
                
                # Common signal mappings
                signal_mappings = {
                    'PV': 'PV',     # Process Value
                    'SP': 'SP',     # Set Point  
                    'OUT': 'OP',    # Output
                    'CV': 'PV',     # Control Value → Process Value
                    'MV': 'OP',     # Manipulated Variable → Output
                }
                
                # Map signal name if needed
                mapped_signal = signal_mappings.get(signal_name.upper(), signal_name)
                
                return f"{instrument_part}.{mapped_signal}"
        
        # Strategy 2: Handle dot-separated paths
        elif '.' in cleaned_path:
            parts = cleaned_path.split('.')
            if len(parts) >= 2:
                # Assume last part is the signal, everything else is instrument
                instrument_part = '.'.join(parts[:-1])
                signal_part = parts[-1]
                return f"{instrument_part}.{signal_part}"
        
        # Strategy 3: If no clear structure, try to extract instrument pattern
        else:
            # Look for instrument-like patterns
            patterns = [
                r'([A-Z]\d{2}[A-Z]{2}\d{4}[A-Z]?)',    # E20FC0023
                r'([A-Z]{2,4}\d{2,4}[A-Z]?)',          # FIC101, TIC23A
                r'([A-Z]{3,8}\d{1,4}[A-Z]?)',          # TANK101, SUAT91D
            ]
            
            for pattern in patterns:
                matches = re.findall(pattern, cleaned_path.upper())
                if matches:
                    return f"{matches[0]}.PV"  # Default to .PV
        
        # Fallback: return as-is
        return cleaned_path
    
    def export_csv(self, file_path):
        """Export to CSV format with embedded metadata headers (comma-delimited)"""
        clean_df = self.get_clean_dataframe()
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=',')
            
            # Row 1: Tag names (column headers)
            tag_row = ['Timestamp'] + [col for col in clean_df.columns if col != 'Timestamp']
            writer.writerow(tag_row)
            
            # Row 2: Descriptions
            desc_row = ['Time'] + [self.descriptions.get(col, '') for col in clean_df.columns if col != 'Timestamp']
            writer.writerow(desc_row)
            
            # Row 3: Units
            units_row = [''] + [self.units.get(col, '') for col in clean_df.columns if col != 'Timestamp']
            writer.writerow(units_row)
            
            # Blank row separator (optional)
            writer.writerow([])
            
            # Data rows - convert DataFrame to CSV manually to maintain control
            for _, row in clean_df.iterrows():
                data_row = [row['Timestamp']] + [row[col] for col in clean_df.columns if col != 'Timestamp']
                writer.writerow(data_row)
    
    def export_tsv(self, file_path):
        """Export to TSV format (tab-delimited, clean data)"""
        clean_df = self.get_clean_dataframe()
        clean_df.to_csv(file_path, sep='\t', index=False)
    
    def export_xlsx(self, file_path):
        """Export to Excel XLSX format with embedded metadata headers"""
        clean_df = self.get_clean_dataframe()
        
        # Create workbook directly using openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'PI Data'
        
        # Prepare data for writing
        columns = ['Timestamp'] + [col for col in clean_df.columns if col != 'Timestamp']
        
        # Row 1: Tag names (column headers)
        for col_idx, col_name in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        # Row 2: Descriptions  
        worksheet.cell(row=2, column=1, value='Time').font = Font(italic=True, color="666666")
        worksheet.cell(row=2, column=1).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        for col_idx, col_name in enumerate([col for col in columns if col != 'Timestamp'], 2):
            description = self.descriptions.get(col_name, '')
            cell = worksheet.cell(row=2, column=col_idx, value=description)
            cell.font = Font(italic=True, color="666666")
            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Row 3: Units
        worksheet.cell(row=3, column=1, value='')  # Empty for timestamp
        for col_idx, col_name in enumerate([col for col in columns if col != 'Timestamp'], 2):
            units = self.units.get(col_name, '')
            cell = worksheet.cell(row=3, column=col_idx, value=units)
            cell.font = Font(size=10, color="888888")
            cell.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        
        # Row 4: Blank separator (skip)
        
        # Data rows starting from row 5
        data_start_row = 5
        for data_row_idx, (_, row) in enumerate(clean_df.iterrows(), data_start_row):
            # Timestamp column
            timestamp_cell = worksheet.cell(row=data_row_idx, column=1, value=row['Timestamp'])
            if pd.api.types.is_datetime64_any_dtype(clean_df['Timestamp']):
                # If it's a datetime, convert to proper Excel datetime
                timestamp_cell.value = pd.to_datetime(row['Timestamp']).to_pydatetime()
                timestamp_cell.number_format = 'MM/DD/YYYY HH:MM:SS'
            
            # Data columns
            for col_idx, col_name in enumerate([col for col in columns if col != 'Timestamp'], 2):
                value = row[col_name]
                # Handle NaN values
                if pd.isna(value):
                    value = None
                worksheet.cell(row=data_row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(file_path)
    
    def export_txt(self, file_path):
        """Export to DMC TXT format with INSTRUMENT TAG REPLACEMENT"""
        df = self.dataframe.copy()
        df['Timestamp'] = pd.to_datetime(df['Timestamp'], utc=True)
        
        # Get all tags (excluding Timestamp and Status column)
        all_tag_columns = [col for col in df.columns 
                        if col != 'Timestamp' and col != 'Status']
        
        available_tags = []
        tag_name_mapping = {}  # Maps original tag name to instrument tag name
        
        for tag in all_tag_columns:
            if tag in df.columns:
                available_tags.append(tag)
                
                # INSTRUMENT TAG REPLACEMENT: Convert OPC path to instrument tag
                if tag in self.instrument_mapping:
                    opc_path = self.instrument_mapping[tag]
                    instrument_tag = self.parse_instrument_tag_from_opc_path(opc_path)
                    if instrument_tag:
                        tag_name_mapping[tag] = instrument_tag
                    else:
                        tag_name_mapping[tag] = tag  # Keep original if parsing fails
                else:
                    tag_name_mapping[tag] = tag  # Keep original if no mapping
        
        # Ensure we have the status column
        if "Status" not in df.columns:
            df["Status"] = 'G'  # Default to Good quality
        
        with open(file_path, "w", newline='') as f:
            writer = csv.writer(f, delimiter='\t')
            
            # Header line with timezone
            writer.writerow([f"(timezone:{self.timezone})"])
            
            # Column headers: Time, then INSTRUMENT_TAG1, Status, INSTRUMENT_TAG2, Status, etc.
            header_row = ["Time"]
            for tag in available_tags:
                instrument_tag_name = tag_name_mapping[tag]
                header_row.extend([instrument_tag_name, "Status"])
            writer.writerow(header_row)
            
            # Description row (use instrument tag names in descriptions)
            desc_row = [""]  # Empty for Time column
            for tag in available_tags:
                tag_desc = self.descriptions.get(tag, '')
                # If we have an instrument mapping, note it in the description
                if tag in self.instrument_mapping and tag_name_mapping[tag] != tag:
                    if tag_desc:
                        tag_desc += f" (Original: {tag})"
                    else:
                        tag_desc = f"Mapped from {tag}"
                desc_row.extend([tag_desc, ""])
            writer.writerow(desc_row)
            
            # Units row  
            units_row = [""]  # Empty for Time column
            for tag in available_tags:
                tag_units = self.units.get(tag, '')
                units_row.extend([tag_units, ""])
            writer.writerow(units_row)
            
            # Data rows
            for _, row in df.iterrows():
                timestamp_str = pd.to_datetime(row["Timestamp"]).isoformat()
                row_data = [timestamp_str]
                
                for tag in available_tags:
                    value = row.get(tag, '')
                    if pd.isna(value):
                        value = ''
                    
                    status = row.get("Status", 'G')
                    if pd.isna(status) or status == '':
                        status = 'G'
                    
                    row_data.extend([value, status])
                
                writer.writerow(row_data)
    
    def export_iq(self, file_path):
        """Export to IQ format (tab-delimited, lab data compatible, MM/DD/YYYY format)"""
        clean_df = self.get_clean_dataframe()
        
        # Format timestamp to match lab data format (MM/DD/YYYY HH:MM:SS)
        if "Timestamp" in clean_df.columns:
            clean_df["Timestamp"] = pd.to_datetime(clean_df["Timestamp"]).dt.strftime("%m/%d/%Y %H:%M:%S")
            # Rename to "Time" to match lab data format
            clean_df = clean_df.rename(columns={"Timestamp": "Time"})
        
        # Export with tab delimiter, no index, no quotes
        clean_df.to_csv(file_path, sep='\t', index=False, quoting=csv.QUOTE_NONE, escapechar='\\')